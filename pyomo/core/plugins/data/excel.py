#  ___________________________________________________________________________
#
#  Pyomo: Python Optimization Modeling Objects
#  Copyright 2017 National Technology and Engineering Solutions of Sandia, LLC
#  Under the terms of Contract DE-NA0003525 with National Technology and
#  Engineering Solutions of Sandia, LLC, the U.S. Government retains certain
#  rights in this software.
#  This software is distributed under the 3-clause BSD License.
#  ___________________________________________________________________________


from six import iteritems, string_types
from six.moves import xrange

import pandas as pd

from pyomo.core.base import (Constraint, Param, Set, RangeSet, Var, Objective,
                             Block, Suffix, Expression, value, SortComponents)
from pyomo.gdp import Disjunct


scalar_types = set(['Param', 'Var', 'Expression', 'Constraint', 'Objective'])


def write_dataframes(model):
    valid_ctypes = set([Var, Constraint, Param, Set, RangeSet,
                        Expression, Objective, Block, Disjunct])
    objects = []
    scalars = dict()
    for typ in scalar_types:
        scalars[typ] = []
    for obj in model.component_objects(active=True, descend_into=False,
                                       sort=SortComponents.alphaOrder):
        if obj.type() in valid_ctypes:
            if obj.name == 'TOC':
                raise ValueError("Cannot have component named TOC. This fact "
                                 "is subject to change.")
            objects.append(obj)
        if obj.type().__name__ in scalars and obj.dim() == 0:
            scalars[obj.type().__name__].append(obj)


    # Table of Contents
    toc_data = {'Name'  : [obj.name for obj in objects],
                'Type'  : [obj.type().__name__ for obj in objects],
                'Dim'   : [obj.dim() for obj in objects],
                'Count' : [len(obj) for obj in objects],
                'Doc'   : [obj.doc for obj in objects]}
    TOC = (pd.DataFrame(toc_data)
        .set_index('Name')[['Type', 'Dim', 'Count', 'Doc']])
    # Store model/block name in TOC dataframe
    # Used with default file naming and with blocks
    # Replace brackets, which are illegal sheet name characters
    TOC.name = model.name


    has_dual = has_rc = False
    for suf in model.component_data_objects(Suffix, active=True):
        if (suf.name == 'dual' and suf.import_enabled()):
            has_dual = True
        elif (suf.name == 'rc' and suf.import_enabled()):
            has_rc = True


    # Set up dataframes for Scalar Sheet
    scalar_frames = []

    # Scalar Params
    param_data = {'Param' : [obj.name for obj in scalars['Param']],
                  'Value' : [obj.value for obj in scalars['Param']],
                  'Doc'   : [obj.doc for obj in scalars['Param']]}
    scalar_frames.append(pd.DataFrame(param_data)
        .set_index('Param')[['Value','Doc']])

    # Scalar Vars
    var_data = {'Var'   : [obj.name for obj in scalars['Var']],
                'Lower' : [obj.lb for obj in scalars['Var']],
                'Value' : [obj.value for obj in scalars['Var']],
                'Upper' : [obj.ub for obj in scalars['Var']],
                'Doc'   : [obj.doc for obj in scalars['Var']]}
    if has_rc:
        var_data['RC'] = []
        for obj in scalars['Var']:
            var_data['RC'].append(model.rc[obj] if obj in model.rc else None)
        scalar_frames.append(pd.DataFrame(var_data)
            .set_index('Var')[['Lower','Value','Upper','RC','Doc']])
    else:
        scalar_frames.append(pd.DataFrame(var_data)
            .set_index('Var')[['Lower','Value','Upper','Doc']])

    # Scalar Expressions
    expr_data = {'Expression' : [obj.name for obj in scalars['Expression']],
                 'Value'      : [try_val(obj) for obj in scalars['Expression']],
                 'Doc'        : [obj.doc for obj in scalars['Expression']]}
    scalar_frames.append(pd.DataFrame(expr_data)
        .set_index('Expression')[['Value','Doc']])

    # Scalar Constraints
    con_data = {'Constraint' : [obj.name for obj in scalars['Constraint']],
                'Lower' : [try_val(obj.lower) for obj in scalars['Constraint']],
                'Body'  : [try_val(obj.body) for obj in scalars['Constraint']],
                'Upper' : [try_val(obj.upper) for obj in scalars['Constraint']],
                'Doc'   : [obj.doc for obj in scalars['Constraint']]}
    if has_dual:
        con_data['Dual'] = []
        for obj in scalars['Constraint']:
            con_data['Dual'].append(model.dual[obj] if obj in model.dual else
                                    None)
        scalar_frames.append(pd.DataFrame(con_data)
            .set_index('Constraint')[['Lower','Body','Upper','Dual','Doc']])
    else:
        scalar_frames.append(pd.DataFrame(con_data)
            .set_index('Constraint')[['Lower','Body','Upper','Doc']])

    # Scalar Objectives
    objctv_data = {'Objective' : [obj.name for obj in scalars['Objective']],
                   'Value'     : [try_val(obj) for obj in scalars['Objective']],
                   'Doc'       : [obj.doc for obj in scalars['Objective']]}
    scalar_frames.append(pd.DataFrame(objctv_data)
        .set_index('Objective')[['Value','Doc']])


    # Set up dataframes for each object
    obj_frames = []

    for obj in objects:
        if obj.type() in (Block, Disjunct) and obj.dim() == 0:
            obj_frames.append(write_dataframes(obj))
            continue

        data = dict()

        if obj.type() in (Block, Disjunct) and obj.dim() > 0:
            if obj.dim() == 1:
                indices = [obj.index_set().name]
            else:
                indices = [ind.name for ind in obj._implicit_subsets]
            for ind in indices:
                data[ind] = []

            data['Block'] = []

            for ind, dobj in sorted(obj.iteritems(), key=lambda item: item[0]):
                for i in range(len(indices)):
                    data[indices[i]].append(ind if obj.dim() == 1 else ind[i])
                data['Block'].append(dobj.name)

            # Indexed blocks are stored as a list containing all of their
            # constituent blocks, in order to keep them grouped together for
            # sheet linking purposes. Use string to denote indexed block.
            df = pd.DataFrame(data).set_index(indices)
            indx_blk = [obj, df]

            for dobj in sorted(obj.itervalues(), key=lambda val: val.index()):
                indx_blk.append(write_dataframes(dobj))

            obj_frames.append(indx_blk)

        elif (obj.type() in (Var, Constraint, Param, Objective, Expression) and
            obj.dim() > 0):
            if obj.dim() == 1:
                indices = [obj.index_set().name]
            else:
                indices = [ind.name for ind in obj._implicit_subsets]
            for ind in indices:
                data[ind] = []

            if obj.type() is Var:
                data['Value'] = []
                data['Lower'] = []
                data['Upper'] = []
                if has_rc:
                    data['RC'] = []
            elif obj.type() is Constraint:
                data['Body'] = []
                data['Lower'] = []
                data['Upper'] = []
                if has_dual:
                    data['Dual'] = []
            elif obj.type() in (Param, Objective):
                data['Value'] = []

            if obj.type() in (Var, Constraint):
                for ind, dobj in sorted(obj.iteritems(),
                                        key=lambda item: item[0]):
                    for i in range(len(indices)):
                        data[indices[i]].append(ind if obj.dim() == 1 else
                                                ind[i])
                    if obj.type() is Var:
                        data['Lower'].append(dobj.lb)
                        data['Value'].append(dobj.value)
                        data['Upper'].append(dobj.ub)
                        if has_rc:
                            if dobj in model.rc:
                                data['RC'].append(model.rc[dobj])
                            else:
                                data['RC'].append(None)
                    else:
                        data['Lower'].append(try_val(dobj.lower))
                        data['Body'].append(try_val(dobj.body))
                        data['Upper'].append(try_val(dobj.upper))
                        if has_dual:
                            if dobj in model.dual:
                                data['Dual'].append(model.dual[dobj])
                            else:
                                data['Dual'].append(None)

            elif obj.type() in (Param, Objective, Expression):
                for ind, val in sorted(obj.iteritems(),
                                       key=lambda item: item[0]):
                    for i in range(len(indices)):
                        data[indices[i]].append(ind if obj.dim() == 1 else
                                                ind[i])
                    data['Value'].append(val if obj.type() is Param else
                                         try_val(val))

            df = pd.DataFrame(data).set_index(indices)

            if obj.type() is Var:
                if has_rc:
                    df = df[['Lower', 'Value', 'Upper', 'RC']]
                else:
                    df = df[['Lower', 'Value', 'Upper']]
            elif obj.type() is Constraint:
                if has_dual:
                    df = df[['Lower', 'Body', 'Upper', 'Dual']]
                else:
                    df = df[['Lower', 'Body', 'Upper']]

            obj_frames.append((obj, df))

        elif obj.type() in (Set, RangeSet) and obj.dim() == 0:
            data['Values'] = sorted(obj.data())
            df = pd.DataFrame(data)
            obj_frames.append((obj, df))

    return [TOC, scalar_frames, obj_frames]


def write_excel(all_frames, filename=None, engine=None):
    TOC = all_frames[0]

    if filename is None:
        if not filename.endswith('.xlsx'):
            raise ValueError("Pyomo excel writer only supports writing "
                             "to a .xlsx file.")
        # Default filename is model name, stored in top-level TOC.name
        filename = TOC.name + '.xlsx'
    writer = pd.ExcelWriter(filename, engine=engine)

    objectMap = dict()
    sheetNameMap = dict()
    sheetmap_name = '^SheetMap^'

    def replace_brackets(s):
        return s.replace('[', '-').replace(']', '')

    def sheet_namer(orig_name):
        if orig_name in sheetNameMap:
            # Only create this label once and store it, so we
            # do not return a different label for the same sheet
            return sheetNameMap[orig_name]
        name = replace_brackets(orig_name)
        if len(name) > 31 or name in ('History', sheetmap_name):
            # History is reserved by Excel. sheetmap_name is used below.
            char = '^'
            num = 1
            suffix = char + str(num)
            new_name = name[:31 - len(suffix)] + suffix
            while new_name in sheetNameMap.values():
                num += 1
                suffix = char + str(num)
                new_name = name[:31 - len(suffix)] + suffix
            name = new_name
        sheetNameMap[orig_name] = name
        return name

    _write_excel(all_frames,
                 writer=writer,
                 objectMap=objectMap,
                 sheet_namer=sheet_namer)

    names = []
    sheets = []
    for obj, sheet_name in iteritems(objectMap):
        if isinstance(obj, string_types):
            # Record names of scalar sheets
            names.append(obj)
            sheets.append(sheet_name)
        elif obj.type() is Var:
            # Only Vars need to be loaded with a value after reading back in
            names.append(obj.name)
            sheets.append(sheet_name)

    sheetmap_df = pd.DataFrame({'Name': names, 'Sheet': sheets})
    sheetmap_df.to_excel(writer, sheet_name=sheetmap_name)
    sheetmap_sheet = writer.sheets[sheetmap_name]

    if writer.engine == 'xlsxwriter':
        sheetmap_sheet.hide()

    elif writer.engine[:8] == 'openpyxl':
        sheetmap_sheet.sheet_state = 'hidden'

    writer.save()

    return filename


def _write_excel(all_frames,
                 writer,
                 objectMap,
                 sheet_namer,
                 parent=None,
                 parent_row=None):

    TOC, scalar_frames, obj_frames = all_frames

    if parent is not None:
        parent_sheet = sheet_namer(parent)

    block = TOC.name


    # Table of Contents
    # One name is the block name, the other is a name safe for sheet naming
    # Different so that link titles show real name with brackets to look nicer
    toc_name = 'TOC' if parent is None else block
    toc_sheet_name = sheet_namer(toc_name)
    TOC.to_excel(writer, sheet_name=toc_sheet_name, merge_cells=False,
                 startrow=0 if parent is None else 1,
                 freeze_panes=(1 + int(bool(parent)), 0))

    # Scalar Sheet
    scalar_startrow = 1
    scalar_name = 'Scalar' if parent is None else \
                  sheet_namer(toc_sheet_name + '.Scalar')
    for df in scalar_frames:
        if df.empty:
            continue
        df.to_excel(writer, sheet_name=scalar_name, startrow=scalar_startrow,
                    merge_cells=False)
        scalar_startrow += len(df.index) + 2

    if scalar_name in writer.sheets:
        # If this block has a scalar sheet, record and map its name.
        objectMap[toc_name] = scalar_name


    if writer.engine == 'xlsxwriter':
        # Create link for each TOC entry to their respective sheet
        toc = writer.sheets[toc_sheet_name]
        link_format = writer.book.add_format({'color'     : 'blue',
                                              'underline' : 1,
                                              'bold'      : 1,
                                              'align'     : 'left'})
        for i in range(len(TOC.index)):
            if TOC['Type'][i] in scalar_types and TOC['Dim'][i] == 0:
                sheet = scalar_name
            else:
                # Use TOC['Name'] -> TOC.index
                sheet = sheet_namer(TOC.index[i])
            # name is what is already in the dataframe: obj.name
            name = TOC.index[i]
            # If there is a parent TOC, move the row down 1 for each entry
            i += int(bool(parent))
            toc.write_url(row=i + 1, col=0, url="internal:'%s'!A1" % sheet,
                          cell_format=link_format, string=name)

        toc.autofilter(int(bool(parent)), 0, int(bool(parent)), 4)

        # Place 'TOC' link on the Scalar sheet
        if scalar_startrow > 1:
            scalar_sheet = writer.sheets[scalar_name]
            scalar_sheet.write_url('A1', "internal:'%s'!A1" % toc_sheet_name,
                                   string=toc_name)

            scalar_sheet.set_column('A:A', 25)
            scalar_sheet.set_column('B:F', 10)

        # Place TOC link on child TOC page
        if parent is not None:
            toc.write_url('A1', string=parent,
                          url="internal:'%s'!A%s" % (parent_sheet, parent_row))

        # Set TOC column widths
        toc.set_column('A:A', 25)
        toc.set_column('B:D', 12)
        toc.set_column('E:E', 20)

    elif writer.engine[:8] == 'openpyxl':
        # Engine is actually 'openpyxl22', only check that it starts
        # with 'openpyxl' in case there might be other versions
        from openpyxl.styles import Font, Color, colors
        from openpyxl.utils import get_column_letter
        from copy import copy

        # Create link for each TOC entry to their respective sheet
        toc = writer.sheets[toc_sheet_name]
        # Use theme=10 to get color to change when link is clicked
        link_color = Color(theme=10)
        name_font = Font(bold=True, underline='single', color=link_color)
        for i in range(len(TOC.index)):
            if TOC['Type'][i] in scalar_types and TOC['Dim'][i] == 0:
                sheet = scalar_name
            else:
                sheet = sheet_namer(TOC.index[i])
            # If there is a parent TOC, move the row down 1 for each entry
            i += int(bool(parent))
            cell = toc.cell(row=i + 2, column=1)
            cell.hyperlink = "#'%s'!A1" % sheet
            cell.font = name_font
            cell.alignment = cell.alignment.copy(horizontal='left')

        toc.auto_filter.ref = "A%s:E%s" % (1 + int(bool(parent)),
                                           1 + int(bool(parent)))

        # Place 'TOC' link on the Scalar sheet
        toc_font = copy(name_font)
        toc_font.bold = False
        if scalar_startrow > 1:
            scalar_sheet = writer.sheets[scalar_name]
            scalar_sheet['A1'].value = toc_name
            scalar_sheet['A1'].hyperlink = "#'%s'!A1" % toc_sheet_name
            scalar_sheet['A1'].font = toc_font

            scalar_sheet.column_dimensions['A'].width = 25
            scalar_sheet.column_dimensions['B'].width = \
            scalar_sheet.column_dimensions['C'].width = \
            scalar_sheet.column_dimensions['D'].width = \
            scalar_sheet.column_dimensions['E'].width = \
            scalar_sheet.column_dimensions['F'].width = 12

        # Place TOC link on child TOC page
        if parent is not None:
            toc['A1'].value = parent
            toc['A1'].hyperlink = "#'%s'!A%s" % (parent_sheet, parent_row)
            toc['A1'].font = toc_font

        # Set TOC column widths
        toc.column_dimensions['A'].width = 25
        toc.column_dimensions['B'].width = \
        toc.column_dimensions['C'].width = \
        toc.column_dimensions['D'].width = 12
        toc.column_dimensions['E'].width = 20


    for item in obj_frames:

        if type(item) is list and type(item[0]) is pd.DataFrame:
            # This is a singleton block's list of dataframes
            # The first item in the list if the TOC dataframe
            child_toc = item[0]
            # Pass row of this block's entry in this TOC so that the child
            # block's TOC can link back to its row in the parent TOC
            toc_row = TOC.index.get_loc(child_toc.name) + 2 + int(bool(parent))
            _write_excel(item, writer=writer, parent_row=toc_row,
                         parent=toc_name, objectMap=objectMap,
                         sheet_namer=sheet_namer)
            # This is the only case where the extra info at
            # the top of the sheet is not needed.
            continue

        if type(item) is list:
            # This is an indexed block, whose first item is the object
            # List items are: object, dataframe listing all subblocks,
            # and then each subblock in the form of a list.
            obj, df = item[0], item[1]
            obj_sheet_name = sheet_namer(obj.name)

            if obj_sheet_name in writer.sheets:
                raise ValueError("Pyomo excel writer requires all components "
                                 "have unique names. Found component: %s twice."
                                 % obj.name)

            df.to_excel(writer, sheet_name=obj_sheet_name, startrow=2,
                        merge_cells=False,
                        freeze_panes=(3, 0))

            col = df.index.nlevels
            i = 0
            for blk in item[2:]:
                child_toc = blk[0]
                blk_name = child_toc.name
                blk_sheet_name = sheet_namer(blk_name)
                parent_row = i + 4 + int(bool(parent))
                _write_excel(blk, writer=writer, parent_row=parent_row,
                             parent=obj.name, objectMap=objectMap,
                             sheet_namer=sheet_namer)

                # Create link for each subblock
                if writer.engine == 'xlsxwriter':
                    sheet = writer.sheets[obj_sheet_name]
                    sheet.write_url(row=i + 3, col=col, string=blk_name,
                                    url="internal:'%s'!A1" % blk_sheet_name)

                elif writer.engine[:8] == 'openpyxl':
                    sheet = writer.sheets[obj_sheet_name]
                    cell = sheet.cell(row=i + 4, column=col + 1)
                    cell.hyperlink = "#'%s'!A1" % blk_sheet_name
                    cell.font = toc_font
                i += 1

        else:
            # type is tuple -> regular object and dataframe
            obj, df = item
            obj_sheet_name = objectMap.setdefault(obj, sheet_namer(obj.name))

            if obj_sheet_name in writer.sheets:
                raise ValueError("Pyomo excel writer requires all components "
                                 "have unique names. Found component: %s twice."
                                 % obj.name)

            # When cells are merged, things like the autofilter do not
            # work. For example the autofilter only recognizes the first row
            # of the merged block when filtering.
            df.to_excel(writer, sheet_name=obj_sheet_name, startrow=2,
                        merge_cells=False,
                        freeze_panes=(3, 0))

        cols = df.index.nlevels + len(df.columns)
        toc_row = TOC.index.get_loc(obj.name) + 2 + int(bool(parent))

        if writer.engine == 'xlsxwriter':
            ws = writer.sheets[obj_sheet_name]

            ws.write_url('A1', "internal:'%s'!A%s" % (toc_sheet_name, toc_row),
                         string=toc_name)
            ws.write('A2', obj.name)
            ws.write('B2', obj.type().__name__)
            ws.write('C2', obj.doc)

            ws.autofilter(2, 0, 2, cols - 1)

            ws.set_column(0, df.index.nlevels, 18)
            ws.set_column(df.index.nlevels, cols, 12)

        elif writer.engine[:8] == 'openpyxl':
            ws = writer.sheets[obj_sheet_name]

            toc_row = TOC.index.get_loc(obj.name) + 2 + int(bool(parent))

            ws['A1'].value = toc_name
            ws['A1'].hyperlink = "#'%s'!A%s" % (toc_sheet_name, toc_row)
            ws['A1'].font = toc_font

            ws['A2'] = obj.name
            ws['B2'] = obj.type().__name__
            ws['C2'] = obj.doc

            ws.auto_filter.ref = "A3:%s3" % get_column_letter(cols)

            for i in xrange(df.index.nlevels):
                ws.column_dimensions[get_column_letter(i + 1)].width = 18
            for i in xrange(df.index.nlevels, cols):
                ws.column_dimensions[get_column_letter(i + 1)].width = 12


def try_val(expr):
    """
    A ValueError is thrown when an expression contains uninitialized components.
    Return None instead so the Excel entry is blank.
    """
    return value(expr, exception=False)




"""
TODO:
Indexed Sets, RangeSets
active column
"""
